import io
from datetime import datetime, time, timedelta
from typing import Optional, Dict, Any, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.employee import Employee, EmployeeStatus
from app.models.email_job import EmailJob, EmailJobStatus, EmailType
from app.models.feedback_record import FeedbackRecord, FeedbackStatus


def apply_excel_styling(ws, headers: List[str]):
    """Applies consistent executive styling to an openpyxl worksheet."""
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    ws.append(headers)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border


def auto_fit_columns(ws):
    """Calculates max character width per column and applies auto-fitting."""
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            if isinstance(cell.value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm"

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)


def parse_date_range_bounds(
    date_filter: Optional[str],
    start_date: Optional[str],
    end_date: Optional[str],
) -> tuple[Optional[datetime], Optional[datetime]]:
    """
    Parses inclusive date-range parameters into timezone-safe UTC start/end datetime boundaries.
    Start boundary: >= start_datetime (00:00:00)
    End boundary: < end_datetime (start of next day)
    """
    if not date_filter and not (start_date and end_date):
        return None, None

    today = datetime.utcnow().date()
    d_start, d_end = None, None

    if date_filter == "today":
        d_start, d_end = today, today
    elif date_filter == "last_7_days":
        d_start, d_end = today - timedelta(days=7), today
    elif date_filter == "last_30_days":
        d_start, d_end = today - timedelta(days=30), today
    elif date_filter == "this_month":
        d_start = today.replace(day=1)
        d_end = today
    elif date_filter == "prev_month":
        first_this_month = today.replace(day=1)
        last_prev_month = first_this_month - timedelta(days=1)
        d_start = last_prev_month.replace(day=1)
        d_end = last_prev_month
    elif date_filter == "custom" or (start_date and end_date):
        if start_date:
            d_start = datetime.strptime(start_date[:10], "%Y-%m-%d").date()
        if end_date:
            d_end = datetime.strptime(end_date[:10], "%Y-%m-%d").date()

    dt_start = datetime.combine(d_start, time(0, 0, 0)) if d_start else None
    dt_end = datetime.combine(d_end + timedelta(days=1), time(0, 0, 0)) if d_end else None

    return dt_start, dt_end


def generate_employee_report(
    db: Session,
    date_filter: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    emp_status: Optional[str] = None,
    feedback_status: Optional[str] = None,
) -> bytes:
    """Generates comprehensive Employee Exit Feedback Report."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Feedback Report"

    headers = [
        "Full Name",
        "Personal Email",
        "Last Working Date",
        "Feedback Due Date",
        "Employee Status",
        "Feedback Status",
        "Initial Email Status",
        "Initial Email Sent At",
        "Reminder 1 Status",
        "Reminder 1 Sent At",
        "Reminder 2 Status",
        "Reminder 2 Sent At",
        "Latest Reminder Status",
        "Feedback Submitted At",
    ]
    apply_excel_styling(ws, headers)

    dt_start, dt_end = parse_date_range_bounds(date_filter, start_date, end_date)

    query = db.query(Employee)
    if emp_status:
        query = query.filter(Employee.status == emp_status.strip().upper())

    if dt_start and dt_end:
        query = query.filter(
            Employee.feedback_due_date >= dt_start.date(),
            Employee.feedback_due_date < dt_end.date(),
        )

    employees = query.order_by(Employee.feedback_due_date.desc()).all()

    for emp in employees:
        fb_rec = db.query(FeedbackRecord).filter(FeedbackRecord.employee_id == emp.id).first()
        fb_stat = fb_rec.status if fb_rec else "PENDING"
        if feedback_status and fb_stat != feedback_status.strip().upper():
            continue

        initial_job = db.query(EmailJob).filter(
            EmailJob.employee_id == emp.id,
            EmailJob.email_type == EmailType.EXIT_FEEDBACK_INITIAL,
        ).first()

        r1_job = db.query(EmailJob).filter(
            EmailJob.employee_id == emp.id,
            EmailJob.email_type == EmailType.EXIT_FEEDBACK_REMINDER_1,
        ).first()

        r2_job = db.query(EmailJob).filter(
            EmailJob.employee_id == emp.id,
            EmailJob.email_type == EmailType.EXIT_FEEDBACK_REMINDER_2,
        ).first()

        latest_r_job = db.query(EmailJob).filter(
            EmailJob.employee_id == emp.id,
            EmailJob.email_type.like("EXIT_FEEDBACK_REMINDER_%"),
        ).order_by(EmailJob.scheduled_at.desc()).first()

        ws.append([
            emp.full_name,
            emp.personal_email,
            str(emp.last_working_date),
            str(emp.feedback_due_date),
            emp.status,
            fb_stat,
            initial_job.status if initial_job else "NOT_SCHEDULED",
            initial_job.sent_at.strftime("%Y-%m-%d %H:%M") if initial_job and initial_job.sent_at else "—",
            r1_job.status if r1_job else "NOT_SCHEDULED",
            r1_job.sent_at.strftime("%Y-%m-%d %H:%M") if r1_job and r1_job.sent_at else "—",
            r2_job.status if r2_job else "NOT_SCHEDULED",
            r2_job.sent_at.strftime("%Y-%m-%d %H:%M") if r2_job and r2_job.sent_at else "—",
            latest_r_job.status if latest_r_job else "NOT_SCHEDULED",
            fb_rec.submitted_at.strftime("%Y-%m-%d %H:%M") if fb_rec and fb_rec.submitted_at else "—",
        ])

    auto_fit_columns(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_feedback_report(
    db: Session,
    date_filter: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> bytes:
    """Generates Feedback Focused Report."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feedback Lifecycle Report"

    headers = [
        "Full Name",
        "Last Working Date",
        "Feedback Due Date",
        "Feedback Status",
        "Initial Email Sent At",
        "Reminder 1 Status",
        "Reminder 2 Status",
        "Feedback Submitted At",
        "Feedback Expiry Date",
    ]
    apply_excel_styling(ws, headers)

    dt_start, dt_end = parse_date_range_bounds(date_filter, start_date, end_date)

    query = db.query(Employee)
    if dt_start and dt_end:
        query = query.filter(
            Employee.feedback_due_date >= dt_start.date(),
            Employee.feedback_due_date < dt_end.date(),
        )

    employees = query.order_by(Employee.feedback_due_date.desc()).all()

    for emp in employees:
        fb_rec = db.query(FeedbackRecord).filter(FeedbackRecord.employee_id == emp.id).first()
        initial_job = db.query(EmailJob).filter(
            EmailJob.employee_id == emp.id,
            EmailJob.email_type == EmailType.EXIT_FEEDBACK_INITIAL,
        ).first()

        r1_job = db.query(EmailJob).filter(
            EmailJob.employee_id == emp.id,
            EmailJob.email_type == EmailType.EXIT_FEEDBACK_REMINDER_1,
        ).first()

        r2_job = db.query(EmailJob).filter(
            EmailJob.employee_id == emp.id,
            EmailJob.email_type == EmailType.EXIT_FEEDBACK_REMINDER_2,
        ).first()

        ws.append([
            emp.full_name,
            str(emp.last_working_date),
            str(emp.feedback_due_date),
            fb_rec.status if fb_rec else "PENDING",
            initial_job.sent_at.strftime("%Y-%m-%d %H:%M") if initial_job and initial_job.sent_at else "—",
            r1_job.status if r1_job else "NOT_SCHEDULED",
            r2_job.status if r2_job else "NOT_SCHEDULED",
            fb_rec.submitted_at.strftime("%Y-%m-%d %H:%M") if fb_rec and fb_rec.submitted_at else "—",
            fb_rec.expires_at.strftime("%Y-%m-%d %H:%M") if fb_rec and fb_rec.expires_at else "—",
        ])

    auto_fit_columns(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_email_report(
    db: Session,
    date_filter: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    email_type: Optional[str] = None,
    job_status: Optional[str] = None,
) -> bytes:
    """Generates Email Dispatch History Report."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Email Delivery Report"

    headers = [
        "Full Name",
        "Recipient Email",
        "Email Type",
        "Job Status",
        "Scheduled At (UTC)",
        "Sent At (UTC)",
        "Attempt Count",
        "Last Error",
    ]
    apply_excel_styling(ws, headers)

    dt_start, dt_end = parse_date_range_bounds(date_filter, start_date, end_date)

    query = db.query(EmailJob, Employee).join(Employee, EmailJob.employee_id == Employee.id)

    if email_type:
        query = query.filter(EmailJob.email_type == email_type.strip().upper())
    if job_status:
        query = query.filter(EmailJob.status == job_status.strip().upper())

    if dt_start and dt_end:
        query = query.filter(
            EmailJob.scheduled_at >= dt_start,
            EmailJob.scheduled_at < dt_end,
        )

    results = query.order_by(EmailJob.scheduled_at.desc()).all()

    for job, emp in results:
        ws.append([
            emp.full_name,
            job.recipient_email,
            job.email_type,
            job.status,
            job.scheduled_at.strftime("%Y-%m-%d %H:%M") if job.scheduled_at else "—",
            job.sent_at.strftime("%Y-%m-%d %H:%M") if job.sent_at else "—",
            job.attempt_count,
            job.last_error or "—",
        ])

    auto_fit_columns(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
