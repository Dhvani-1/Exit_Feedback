import io
import openpyxl
import pytest
from datetime import datetime, timedelta
from app.models.employee import Employee, EmployeeStatus


def create_mock_excel(rows):
    """Utility to generate in-memory Excel bytes from rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_excel_template_download(auth_client):
    """Verifies downloading standard 3-column Excel template."""
    res = auth_client.get("/api/employees/excel-template")
    assert res.status_code == 200
    assert "application/vnd.openxmlformats-officedocument" in res.headers["content-type"]
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == ("Full Name", "Personal Email", "Last Working Date", "Designation", "Start Date", "Tenure")


def test_excel_upload_preview_valid_and_invalid(auth_client, db):
    """Tests upload preview validation with valid, invalid, duplicate, and empty rows."""
    excel_bytes = create_mock_excel([
        [" Full Name ", " Personal Email ", " Last Working Date "],
        ["Rahul Sharma", "rahul@gmail.com", "15/01/2027"],
        ["Priya Shah", "priya@gmail.com", "2027-02-28"],
        ["", "", ""],  # empty row, ignored
        ["Invalid Email Person", "not-an-email", "15/01/2027"],  # invalid email
        ["Invalid Date Person", "dateerr@gmail.com", "not-a-date"],  # invalid date
        ["Rahul Sharma", "rahul@gmail.com", "15/01/2027"],  # duplicate row in file
    ])

    files = {"file": ("test.xlsx", io.BytesIO(excel_bytes), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    res = auth_client.post("/api/employees/upload-preview", files=files)

    assert res.status_code == 200
    data = res.json()
    assert data["total_rows"] == 5
    assert data["valid_count"] == 2
    assert data["invalid_count"] == 3


def test_excel_confirm_import_and_duplicate_prevention(auth_client, db):
    """Tests confirming import and preventing duplicates on email + last_working_date."""
    valid_rows = [
        {
            "row_number": 2,
            "full_name": "Amit Patel",
            "personal_email": "amit@gmail.com",
            "last_working_date": "2027-03-31",
            "is_duplicate": False,
        },
        {
            "row_number": 3,
            "full_name": "Sita Verma",
            "personal_email": "sita@gmail.com",
            "last_working_date": "2027-04-15",
            "is_duplicate": False,
        },
    ]

    res = auth_client.post("/api/employees/import-confirm", json={"valid_rows": valid_rows})
    assert res.status_code == 200
    assert res.json()["imported_count"] == 2

    # Verify DB records created
    emp1 = db.query(Employee).filter(Employee.personal_email == "amit@gmail.com").first()
    assert emp1 is not None
    assert emp1.full_name == "Amit Patel"
    assert str(emp1.last_working_date) == "2027-03-31"

    # Attempt re-import of identical payload (should skip duplicates)
    res2 = auth_client.post("/api/employees/import-confirm", json={"valid_rows": valid_rows})
    assert res2.status_code == 200
    assert res2.json()["imported_count"] == 0
    assert res2.json()["skipped_duplicates_count"] == 2


def test_excel_same_email_different_lwd_allowed(auth_client, db):
    """Verifies that a person with the same personal email but different Last Working Date is allowed (rehire)."""
    # 1. Create first record
    emp1 = Employee(
        full_name="Rehire Person",
        personal_email="rehire@example.com",
        last_working_date=datetime.strptime("2025-01-15", "%Y-%m-%d").date(),
        feedback_due_date=datetime.strptime("2025-04-15", "%Y-%m-%d").date(),
        status=EmployeeStatus.SCHEDULED,
    )
    db.add(emp1)
    db.commit()

    # 2. Import second record with same email but new LWD
    valid_rows = [
        {
            "row_number": 2,
            "full_name": "Rehire Person",
            "personal_email": "rehire@example.com",
            "last_working_date": "2027-06-30",
            "is_duplicate": False,
        }
    ]

    res = auth_client.post("/api/employees/import-confirm", json={"valid_rows": valid_rows})
    assert res.status_code == 200
    assert res.json()["imported_count"] == 1

    records = db.query(Employee).filter(Employee.personal_email == "rehire@example.com").all()
    assert len(records) == 2


def test_excel_error_report_export(auth_client):
    """Verifies generating downloadable Excel error report for invalid rows."""
    invalid_rows = [
        {
            "row_number": 14,
            "full_name": "John Doe",
            "personal_email": "invalid-email",
            "last_working_date": "15/01/2027",
            "error_reason": "Invalid email address",
        }
    ]

    res = auth_client.post("/api/employees/export-error-report", json={"invalid_rows": invalid_rows})
    assert res.status_code == 200
    assert "application/vnd.openxmlformats-officedocument" in res.headers["content-type"]
