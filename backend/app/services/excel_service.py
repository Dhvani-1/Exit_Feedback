import re
import io
import openpyxl
from datetime import datetime, date
from typing import List, Dict, Any, Tuple
from pydantic import EmailStr, validate_email
from sqlalchemy.orm import Session

from app.models.employee import Employee
from app.services.date_calculator import calculate_feedback_due_date
from app.services.reminder_service import ensure_feedback_record


def parse_date_value(val: Any) -> Tuple[bool, Any, str]:
    """
    Parses native Excel date values or string date representations into date objects.
    Supported formats:
    - Native datetime or date object
    - DD/MM/YYYY or DD-MM-YYYY (e.g. 15/01/2027, 15-01-2027)
    - YYYY-MM-DD (e.g. 2027-01-15)
    - YYYY/MM/DD
    """
    if val is None:
        return False, None, "Missing Last Working Date"

    if isinstance(val, (datetime, date)):
        parsed = val.date() if isinstance(val, datetime) else val
        return True, parsed, ""

    if isinstance(val, (int, float)):
        # Native Excel float timestamp
        try:
            parsed_dt = openpyxl.utils.datetime.from_excel(val)
            return True, parsed_dt.date(), ""
        except Exception:
            return False, None, "Invalid native Excel date value"

    s_val = str(val).strip()
    if not s_val:
        return False, None, "Missing Last Working Date"

    date_formats = [
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d/%m/%y",
        "%d-%m-%y",
    ]

    for fmt in date_formats:
        try:
            dt = datetime.strptime(s_val, fmt)
            return True, dt.date(), ""
        except ValueError:
            continue

    return False, None, f"Invalid date format '{s_val}'. Use DD/MM/YYYY or YYYY-MM-DD."


def parse_and_validate_excel(file_bytes: bytes, db: Session) -> Dict[str, Any]:
    """
    Parses and validates an uploaded Excel file.
    Normalizes headers, checks for required columns, ignores empty rows,
    validates emails and dates, and identifies duplicate rows within file and DB.
    """
    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return {
            "error_code": "INVALID_FILE",
            "message": f"Could not read Excel file: {str(e)}",
        }

    sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows or len(rows) < 1:
        return {
            "error_code": "EMPTY_FILE",
            "message": "Uploaded Excel file contains no data.",
        }

    # Header normalization
    raw_headers = rows[0]
    headers = [str(h).strip() if h is not None else "" for h in raw_headers]

    required_columns = ["Full Name", "Personal Email", "Last Working Date"]
    col_idx = {}

    for req in required_columns:
        found = False
        for idx, h in enumerate(headers):
            if h.lower() == req.lower():
                col_idx[req] = idx
                found = True
                break
        if not found:
            return {
                "error_code": "MISSING_COLUMNS",
                "message": f"Missing required column header '{req}'. Required headers: Full Name, Personal Email, Last Working Date.",
            }

    optional_columns = {
        "designation": ["Designation", "Role", "Job Title"],
        "start_date": ["Start Date", "Joining Date", "Date of Joining"],
        "tenure": ["Tenure"],
    }
    opt_idx = {}
    for key, aliases in optional_columns.items():
        for alias in aliases:
            for idx, h in enumerate(headers):
                if h.lower() == alias.lower():
                    opt_idx[key] = idx
                    break
            if key in opt_idx:
                break

    valid_rows = []
    invalid_rows = []
    seen_in_file = set()

    for row_num, row in enumerate(rows[1:], start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue  # Ignore completely empty rows

        name_val = row[col_idx["Full Name"]] if col_idx["Full Name"] < len(row) else None
        email_val = row[col_idx["Personal Email"]] if col_idx["Personal Email"] < len(row) else None
        date_val = row[col_idx["Last Working Date"]] if col_idx["Last Working Date"] < len(row) else None

        desig_val = row[opt_idx["designation"]] if "designation" in opt_idx and opt_idx["designation"] < len(row) else None
        start_date_val = row[opt_idx["start_date"]] if "start_date" in opt_idx and opt_idx["start_date"] < len(row) else None
        tenure_val = row[opt_idx["tenure"]] if "tenure" in opt_idx and opt_idx["tenure"] < len(row) else None

        full_name = str(name_val).strip() if name_val is not None else ""
        personal_email = str(email_val).strip().lower() if email_val is not None else ""
        designation = str(desig_val).strip() if desig_val is not None and str(desig_val).strip() != "" else None
        tenure = str(tenure_val).strip() if tenure_val is not None and str(tenure_val).strip() != "" else None

        parsed_start_date = None
        if start_date_val is not None and str(start_date_val).strip() != "":
            is_sd_valid, sd_dt, _ = parse_date_value(start_date_val)
            if is_sd_valid:
                parsed_start_date = sd_dt.isoformat()

        errors = []

        # 1. Full Name check
        if not full_name:
            errors.append("Missing Full Name")

        # 2. Personal Email check
        if not personal_email:
            errors.append("Missing Personal Email")
        else:
            try:
                validate_email(personal_email)
            except Exception:
                errors.append("Invalid email address")

        # 3. Last Working Date check
        is_date_valid, parsed_date, date_err = parse_date_value(date_val)
        if not is_date_valid:
            errors.append(date_err)

        if errors:
            invalid_rows.append({
                "row_number": row_num,
                "full_name": full_name,
                "personal_email": personal_email,
                "last_working_date": str(date_val) if date_val else "",
                "error_reason": ", ".join(errors),
            })
            continue

        # 4. In-file Duplicate Check
        file_key = (personal_email, parsed_date)
        if file_key in seen_in_file:
            invalid_rows.append({
                "row_number": row_num,
                "full_name": full_name,
                "personal_email": personal_email,
                "last_working_date": parsed_date.isoformat(),
                "error_reason": "Duplicate record in uploaded Excel file",
            })
            continue
        seen_in_file.add(file_key)

        # 5. Database Duplicate Check
        existing = db.query(Employee).filter(
            Employee.personal_email == personal_email,
            Employee.last_working_date == parsed_date,
        ).first()

        is_db_duplicate = existing is not None

        due_date = calculate_feedback_due_date(parsed_date)

        valid_rows.append({
            "row_number": row_num,
            "full_name": full_name,
            "personal_email": personal_email,
            "last_working_date": parsed_date.isoformat(),
            "designation": designation,
            "start_date": parsed_start_date,
            "tenure": tenure,
            "feedback_due_date": due_date.isoformat(),
            "is_duplicate": is_db_duplicate,
        })

    total_rows = len(valid_rows) + len(invalid_rows)
    duplicate_count = sum(1 for r in valid_rows if r["is_duplicate"])

    return {
        "total_rows": total_rows,
        "valid_count": len(valid_rows),
        "invalid_count": len(invalid_rows),
        "duplicate_count": duplicate_count,
        "valid_rows": valid_rows,
        "invalid_rows": invalid_rows,
    }


def generate_excel_template() -> bytes:
    """Generates the downloadable Excel Template containing required headers and optional Designation, Start Date, Tenure columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Import Template"

    headers = ["Full Name", "Personal Email", "Last Working Date", "Designation", "Start Date", "Tenure"]
    ws.append(headers)

    example_row = ["Rahul Sharma (Example)", "rahul.example@gmail.com", "15/01/2027", "Senior Engineer", "15/01/2024", "3 years"]
    ws.append(example_row)

    # Column formatting widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 18

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_error_report(invalid_rows: List[Dict[str, Any]]) -> bytes:
    """Generates downloadable Excel Error Report for invalid rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Error Report"

    headers = ["Row Number", "Full Name", "Personal Email", "Last Working Date", "Error Reason"]
    ws.append(headers)

    for item in invalid_rows:
        ws.append([
            item.get("row_number", ""),
            item.get("full_name", ""),
            item.get("personal_email", ""),
            item.get("last_working_date", ""),
            item.get("error_reason", ""),
        ])

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 40

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
